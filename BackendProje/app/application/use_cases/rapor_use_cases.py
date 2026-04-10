"""
Rapor Use Case'leri.

Şablon CRUD, Log yönetimi, Schedule CRUD, Veri sorgulama,
Export ve Schedule tetikleme use case'leri.
"""

from __future__ import annotations
from datetime import date, datetime
from typing import Callable, Dict, List, Optional
import csv
import io

from app.core.repositories.rapor_repository import (
    IRaporSablonuRepository,
    IRaporLoguRepository,
    IRaporScheduleRepository,
    IRaporVeriRepository,
)
from app.core.repositories.sistem_log_repository import ISistemLogRepository
from app.core.entities.rapor import RaporSablonu, RaporLogu, RaporSchedule
from app.core.entities.sistem_log import SistemLog, IslemTipi
from app.core.exceptions import KayitBulunamadiError, GecersizIslemError
from app.application.dto.rapor_dto import (
    RaporSablonuOlusturRequestDTO,
    RaporSablonuGuncelleRequestDTO,
    RaporSablonuResponseDTO,
    RaporLoguResponseDTO,
    RaporScheduleOlusturRequestDTO,
    RaporScheduleGuncelleRequestDTO,
    RaporScheduleResponseDTO,
)


def _rapor_logu_yaz(
    rapor_log_repo: IRaporLoguRepository,
    kullanici_id: int,
    parametreler: dict,
    sablon_id: Optional[int] = None,
) -> None:
    """Rapor logu oluşturur. Tüm veri use case'leri ve export tarafından kullanılır."""
    log = RaporLogu(
        sablon_id=sablon_id,
        kullanici_id=kullanici_id,
        parametreler=parametreler,
    )
    log.basarili_isaretle()
    rapor_log_repo.olustur(log)


class RaporSablonuListeleUseCase:

    def __init__(self, sablon_repo: IRaporSablonuRepository):
        self._sablon_repo = sablon_repo

    def execute(
        self, skip: int = 0, limit: int = 100, tur: Optional[str] = None,
    ) -> List[RaporSablonuResponseDTO]:
        entities = self._sablon_repo.getir_hepsi(skip=skip, limit=limit, tur=tur)
        return [RaporSablonuResponseDTO.from_entity(e) for e in entities]


class RaporSablonuGetirUseCase:

    def __init__(self, sablon_repo: IRaporSablonuRepository):
        self._sablon_repo = sablon_repo

    def execute(self, sablon_id: int) -> RaporSablonuResponseDTO:
        entity = self._sablon_repo.getir_id_ile(sablon_id)
        if not entity:
            raise KayitBulunamadiError("Rapor şablonu", sablon_id)
        return RaporSablonuResponseDTO.from_entity(entity)


class RaporSablonuOlusturUseCase:

    def __init__(
        self, sablon_repo: IRaporSablonuRepository,
        log_repo: ISistemLogRepository,
    ):
        self._sablon_repo = sablon_repo
        self._log_repo = log_repo

    def execute(
        self, dto: RaporSablonuOlusturRequestDTO, kullanici_id: int,
    ) -> RaporSablonuResponseDTO:
        entity = RaporSablonu(
            ad=dto.ad,
            tur=dto.tur,
            aciklama=dto.aciklama or "",
            config=dto.config,
            olusturan_kullanici_id=kullanici_id,
        )
        created = self._sablon_repo.olustur(entity)

        self._log_repo.olustur(SistemLog.olustur(
            kullanici_id=kullanici_id,
            islem_tipi=IslemTipi.CREATE,
            modul="Rapor Yönetimi",
            detay=f"Yeni rapor şablonu oluşturuldu: {created.ad} ({created.tur})",
        ))
        return RaporSablonuResponseDTO.from_entity(created)


class RaporSablonuGuncelleUseCase:

    def __init__(
        self, sablon_repo: IRaporSablonuRepository,
        log_repo: ISistemLogRepository,
    ):
        self._sablon_repo = sablon_repo
        self._log_repo = log_repo

    def execute(
        self, sablon_id: int,
        dto: RaporSablonuGuncelleRequestDTO,
        kullanici_id: int,
    ) -> RaporSablonuResponseDTO:
        entity = self._sablon_repo.getir_id_ile(sablon_id)
        if not entity:
            raise KayitBulunamadiError("Rapor şablonu", sablon_id)

        update_data = dto.model_dump(exclude_unset=True)
        for key, value in update_data.items():
            setattr(entity, key, value)

        updated = self._sablon_repo.guncelle(entity)

        self._log_repo.olustur(SistemLog.olustur(
            kullanici_id=kullanici_id,
            islem_tipi=IslemTipi.UPDATE,
            modul="Rapor Yönetimi",
            detay=f"Rapor şablonu güncellendi: {updated.ad}",
        ))
        return RaporSablonuResponseDTO.from_entity(updated)


class RaporSablonuSilUseCase:

    def __init__(
        self, sablon_repo: IRaporSablonuRepository,
        log_repo: ISistemLogRepository,
    ):
        self._sablon_repo = sablon_repo
        self._log_repo = log_repo

    def execute(self, sablon_id: int, kullanici_id: int) -> None:
        entity = self._sablon_repo.getir_id_ile(sablon_id)
        if not entity:
            raise KayitBulunamadiError("Rapor şablonu", sablon_id)

        self._sablon_repo.sil(sablon_id)

        self._log_repo.olustur(SistemLog.olustur(
            kullanici_id=kullanici_id,
            islem_tipi=IslemTipi.DELETE,
            modul="Rapor Yönetimi",
            detay=f"Rapor şablonu pasife alındı: {entity.ad}",
        ))


class RaporLoguListeleUseCase:

    def __init__(self, log_repo: IRaporLoguRepository):
        self._log_repo = log_repo

    def execute(
        self, skip: int = 0, limit: int = 100,
        sablon_id: Optional[int] = None,
    ) -> List[RaporLoguResponseDTO]:
        entities = self._log_repo.getir_hepsi(skip=skip, limit=limit, sablon_id=sablon_id)
        return [RaporLoguResponseDTO.from_entity(e) for e in entities]


class RaporLoguYazUseCase:

    def __init__(self, log_repo: IRaporLoguRepository):
        self._log_repo = log_repo

    def execute(
        self, parametreler: dict, durum: str,
        kullanici_id: int, sablon_id: Optional[int] = None,
    ) -> RaporLoguResponseDTO:
        entity = RaporLogu(
            sablon_id=sablon_id,
            kullanici_id=kullanici_id,
            parametreler=parametreler,
            durum=durum,
            tamamlanma_tarihi=datetime.utcnow(),
        )
        created = self._log_repo.olustur(entity)
        return RaporLoguResponseDTO.from_entity(created)


class RaporScheduleListeleUseCase:

    def __init__(self, schedule_repo: IRaporScheduleRepository):
        self._schedule_repo = schedule_repo

    def execute(
        self, skip: int = 0, limit: int = 100,
    ) -> List[RaporScheduleResponseDTO]:
        entities = self._schedule_repo.getir_hepsi(skip=skip, limit=limit)
        return [RaporScheduleResponseDTO.from_entity(e) for e in entities]


class RaporScheduleGetirUseCase:

    def __init__(self, schedule_repo: IRaporScheduleRepository):
        self._schedule_repo = schedule_repo

    def execute(self, schedule_id: int) -> RaporScheduleResponseDTO:
        entity = self._schedule_repo.getir_id_ile(schedule_id)
        if not entity:
            raise KayitBulunamadiError("Zamanlı rapor", schedule_id)
        return RaporScheduleResponseDTO.from_entity(entity)


class RaporScheduleOlusturUseCase:

    def __init__(
        self, schedule_repo: IRaporScheduleRepository,
        log_repo: ISistemLogRepository,
    ):
        self._schedule_repo = schedule_repo
        self._log_repo = log_repo

    def execute(
        self, dto: RaporScheduleOlusturRequestDTO, kullanici_id: int,
    ) -> RaporScheduleResponseDTO:
        entity = RaporSchedule(
            sablon_id=dto.sablon_id,
            sablon_adi=dto.sablon_adi,
            periyod=dto.periyod,
            saat=dto.saat,
            alici_emailler=dto.alici_emailler,
            format=dto.format or "pdf",
        )
        created = self._schedule_repo.olustur(entity)

        self._log_repo.olustur(SistemLog.olustur(
            kullanici_id=kullanici_id,
            islem_tipi=IslemTipi.CREATE,
            modul="Rapor Zamanlaması",
            detay=f"Yeni zamanlı rapor oluşturuldu: {created.sablon_adi} ({created.periyod})",
        ))
        return RaporScheduleResponseDTO.from_entity(created)


class RaporScheduleGuncelleUseCase:

    def __init__(
        self, schedule_repo: IRaporScheduleRepository,
        log_repo: ISistemLogRepository,
    ):
        self._schedule_repo = schedule_repo
        self._log_repo = log_repo

    def execute(
        self, schedule_id: int,
        dto: RaporScheduleGuncelleRequestDTO,
        kullanici_id: int,
    ) -> RaporScheduleResponseDTO:
        entity = self._schedule_repo.getir_id_ile(schedule_id)
        if not entity:
            raise KayitBulunamadiError("Zamanlı rapor", schedule_id)

        update_data = dto.model_dump(exclude_unset=True)
        for key, value in update_data.items():
            setattr(entity, key, value)

        updated = self._schedule_repo.guncelle(entity)

        self._log_repo.olustur(SistemLog.olustur(
            kullanici_id=kullanici_id,
            islem_tipi=IslemTipi.UPDATE,
            modul="Rapor Zamanlaması",
            detay=f"Zamanlı rapor güncellendi: {updated.sablon_adi}",
        ))
        return RaporScheduleResponseDTO.from_entity(updated)


class RaporScheduleSilUseCase:

    def __init__(
        self, schedule_repo: IRaporScheduleRepository,
        log_repo: ISistemLogRepository,
    ):
        self._schedule_repo = schedule_repo
        self._log_repo = log_repo

    def execute(self, schedule_id: int, kullanici_id: int) -> None:
        entity = self._schedule_repo.getir_id_ile(schedule_id)
        if not entity:
            raise KayitBulunamadiError("Zamanlı rapor", schedule_id)

        self._schedule_repo.sil(schedule_id)

        self._log_repo.olustur(SistemLog.olustur(
            kullanici_id=kullanici_id,
            islem_tipi=IslemTipi.DELETE,
            modul="Rapor Zamanlaması",
            detay=f"Zamanlı rapor silindi: {entity.sablon_adi}",
        ))


class RaporScheduleTetikleUseCase:

    def __init__(
        self, schedule_repo: IRaporScheduleRepository,
        rapor_log_repo: IRaporLoguRepository,
    ):
        self._schedule_repo = schedule_repo
        self._rapor_log_repo = rapor_log_repo

    def execute(self, schedule_id: int, kullanici_id: int) -> dict:
        entity = self._schedule_repo.getir_id_ile(schedule_id)
        if not entity:
            raise KayitBulunamadiError("Zamanlı rapor", schedule_id)

        entity.calistirildi_isaretle()
        self._schedule_repo.guncelle(entity)

        _rapor_logu_yaz(
            self._rapor_log_repo, kullanici_id,
            parametreler={"periyod": entity.periyod, "format": entity.format, "tetikleyen": "manuel"},
            sablon_id=entity.sablon_id,
        )

        return {
            "message": f"'{entity.sablon_adi}' raporu tetiklendi",
            "schedule_id": schedule_id,
            "son_calistirilma": entity.son_calistirilma,
        }


class RaporVeriSorgulaUseCase:
    """Tüm rapor veri sorgularını tek use case üzerinden yönetir."""

    _DISPATCH: Dict[str, Callable] = {
        "stok": lambda repo, **kw: repo.stok_verisi(urun_id=kw.get("urun_id")),
        "siparis": lambda repo, **kw: repo.siparis_verisi(kw.get("baslang_tarihi"), kw.get("bitis_tarihi")),
        "hareket": lambda repo, **kw: repo.hareket_verisi(kw.get("baslang_tarihi"), kw.get("bitis_tarihi")),
        "kritik_stok": lambda repo, **kw: repo.kritik_stok(),
        "skt": lambda repo, **kw: repo.skt_yaklasan(gun=kw.get("gun", 30)),
        "abc": lambda repo, **kw: repo.abc_analiz(),
        "doluluk": lambda repo, **kw: repo.depo_doluluk(),
    }

    def __init__(
        self, veri_repo: IRaporVeriRepository,
        rapor_log_repo: IRaporLoguRepository,
    ):
        self._veri_repo = veri_repo
        self._rapor_log_repo = rapor_log_repo

    def execute(self, tur: str, kullanici_id: int, **kwargs) -> dict:
        fn = self._DISPATCH.get(tur)
        if not fn:
            raise GecersizIslemError(f"Geçersiz rapor türü: {tur}")
        veri = fn(self._veri_repo, **kwargs)
        if tur != "doluluk":
            _rapor_logu_yaz(self._rapor_log_repo, kullanici_id, {"tur": tur, **kwargs})
        return {"veri": veri}


# Backward-compatible aliases — router ve DI container bu isimleri kullanıyor
RaporStokVerisiUseCase = RaporVeriSorgulaUseCase
RaporSiparisVerisiUseCase = RaporVeriSorgulaUseCase
RaporHareketVerisiUseCase = RaporVeriSorgulaUseCase
RaporKritikStokUseCase = RaporVeriSorgulaUseCase
RaporSktUseCase = RaporVeriSorgulaUseCase
RaporAbcAnalizUseCase = RaporVeriSorgulaUseCase
RaporDolulukUseCase = RaporVeriSorgulaUseCase


class RaporExportUseCase:
    """Raporu CSV veya Excel olarak dışa aktar."""

    def __init__(
        self, veri_repo: IRaporVeriRepository,
        rapor_log_repo: IRaporLoguRepository,
    ):
        self._veri_repo = veri_repo
        self._rapor_log_repo = rapor_log_repo

    def execute(
        self, tur: str, format: str,
        kullanici_id: int,
        baslang_tarihi: Optional[date] = None,
        bitis_tarihi: Optional[date] = None,
        sablon_id: Optional[int] = None,
    ) -> tuple:
        """Döner: (bytes_io, media_type, filename)."""
        fn = RaporVeriSorgulaUseCase._DISPATCH.get(tur)
        if not fn:
            raise GecersizIslemError(f"Geçersiz rapor türü: {tur}")

        rows = fn(self._veri_repo, baslang_tarihi=baslang_tarihi, bitis_tarihi=bitis_tarihi)
        if not rows:
            raise KayitBulunamadiError("Rapor verisi", tur)

        _rapor_logu_yaz(
            self._rapor_log_repo, kullanici_id,
            {"tur": tur, "format": format}, sablon_id=sablon_id,
        )

        if format == "csv":
            return self._to_csv(rows, tur)
        if format == "excel":
            return self._to_excel(rows, tur)
        raise GecersizIslemError("Geçersiz format. 'csv' veya 'excel' kullanın.")

    def _to_row_dicts(self, rows) -> List[dict]:
        if hasattr(rows[0], "_mapping"):
            return [dict(r._mapping) for r in rows]
        return rows

    def _to_csv(self, rows, tur: str) -> tuple:
        row_dicts = self._to_row_dicts(rows)
        buffer = io.BytesIO()
        wrapper = io.TextIOWrapper(buffer, encoding="utf-8-sig", newline="")
        writer = csv.DictWriter(wrapper, fieldnames=row_dicts[0].keys())
        writer.writeheader()
        writer.writerows(row_dicts)
        wrapper.detach()
        buffer.seek(0)
        return buffer, "text/csv", f"rapor_{tur}.csv"

    def _to_excel(self, rows, tur: str) -> tuple:
        import openpyxl
        from openpyxl.styles import Font

        row_dicts = self._to_row_dicts(rows)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = tur.capitalize()

        headers = list(row_dicts[0].keys())
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=str(header))
            cell.font = Font(bold=True)

        for row_idx, row_data in enumerate(row_dicts, 2):
            for col_idx, value in enumerate(row_data.values(), 1):
                ws.cell(row=row_idx, column=col_idx, value=str(value) if value is not None else "")

        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        return (
            excel_buffer,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            f"rapor_{tur}.xlsx",
        )
