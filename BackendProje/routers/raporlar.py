from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import Optional
from datetime import date
import io
import csv

from database import get_db
from auth import require_role
from models import Kullanici
from schemas import (
    RaporSablonuCreate, RaporSablonuUpdate, RaporSablonuResponse,
    RaporLoguResponse,
    RaporScheduleCreate, RaporScheduleUpdate, RaporScheduleResponse,
)
from services.rapor_service import RaporService

router = APIRouter(prefix="/api/raporlar", tags=["Raporlar"])


# ========================
# RAPOR ŞABLONLARI
# ========================

@router.get("/sablonlar", response_model=list[RaporSablonuResponse])
def rapor_sablonlarini_listele(
    skip: int = 0,
    limit: int = 100,
    tur: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: Kullanici = Depends(require_role("admin", "lojistik")),
):
    return RaporService.get_sablonlar(db, skip=skip, limit=limit, tur=tur)


@router.get("/sablonlar/{sablon_id}", response_model=RaporSablonuResponse)
def rapor_sablonunu_getir(
    sablon_id: int,
    db: Session = Depends(get_db),
    current_user: Kullanici = Depends(require_role("admin", "lojistik")),
):
    return RaporService.get_sablon(db, sablon_id)


@router.post("/sablonlar", response_model=RaporSablonuResponse, status_code=201)
def rapor_sablonu_ekle(
    sablon: RaporSablonuCreate,
    db: Session = Depends(get_db),
    current_user: Kullanici = Depends(require_role("admin")),
):
    return RaporService.create_sablon(db, sablon, kullanici_id=current_user.id)


@router.put("/sablonlar/{sablon_id}", response_model=RaporSablonuResponse)
def rapor_sablonunu_guncelle(
    sablon_id: int,
    sablon_update: RaporSablonuUpdate,
    db: Session = Depends(get_db),
    current_user: Kullanici = Depends(require_role("admin")),
):
    return RaporService.update_sablon(db, sablon_id, sablon_update, kullanici_id=current_user.id)


@router.delete("/sablonlar/{sablon_id}")
def rapor_sablonunu_sil(
    sablon_id: int,
    db: Session = Depends(get_db),
    current_user: Kullanici = Depends(require_role("admin")),
):
    RaporService.delete_sablon(db, sablon_id, kullanici_id=current_user.id)
    return {"message": "Rapor şablonu başarıyla pasife alındı", "id": sablon_id}


# ========================
# RAPOR VERİLERİ
# ========================

@router.get("/stok/veriler")
def stok_raporu_verisi(
    urun_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    current_user: Kullanici = Depends(require_role("admin", "lojistik", "depocu")),
):
    veri = RaporService.get_stok_verisi(db, urun_id=urun_id)
    RaporService.log_yaz(db, {"urun_id": urun_id, "tur": "stok"}, "Basarili", current_user.id)
    return {"veri": veri}


@router.get("/siparis/veriler")
def siparis_raporu_verisi(
    baslang_tarihi: Optional[date] = Query(None),
    bitis_tarihi: Optional[date] = Query(None),
    db: Session = Depends(get_db),
    current_user: Kullanici = Depends(require_role("admin", "lojistik")),
):
    veri = RaporService.get_siparis_verisi(db, baslang_tarihi, bitis_tarihi)
    RaporService.log_yaz(
        db,
        {"baslang_tarihi": str(baslang_tarihi), "bitis_tarihi": str(bitis_tarihi), "tur": "siparis"},
        "Basarili", current_user.id,
    )
    return {"veri": veri}


@router.get("/hareket/veriler")
def hareket_raporu_verisi(
    baslang_tarihi: Optional[date] = Query(None),
    bitis_tarihi: Optional[date] = Query(None),
    db: Session = Depends(get_db),
    current_user: Kullanici = Depends(require_role("admin", "lojistik", "depocu")),
):
    veri = RaporService.get_hareket_verisi(db, baslang_tarihi, bitis_tarihi)
    RaporService.log_yaz(
        db,
        {"baslang_tarihi": str(baslang_tarihi), "bitis_tarihi": str(bitis_tarihi), "tur": "hareket"},
        "Basarili", current_user.id,
    )
    return {"veri": veri}


@router.get("/kritik-stok")
def kritik_stok_raporu(
    db: Session = Depends(get_db),
    current_user: Kullanici = Depends(require_role("admin", "lojistik", "depocu")),
):
    veri = RaporService.get_kritik_stok(db)
    RaporService.log_yaz(db, {"tur": "kritik_stok"}, "Basarili", current_user.id)
    return {"veri": veri}


@router.get("/skt")
def skt_raporu(
    gun: int = Query(30, description="Kaç gün içinde SKT dolacak"),
    db: Session = Depends(get_db),
    current_user: Kullanici = Depends(require_role("admin", "lojistik", "depocu")),
):
    veri = RaporService.get_skt(db, gun=gun)
    RaporService.log_yaz(db, {"tur": "skt", "gun": gun}, "Basarili", current_user.id)
    return {"veri": veri}


@router.get("/abc-analiz")
def abc_analiz_raporu(
    db: Session = Depends(get_db),
    current_user: Kullanici = Depends(require_role("admin", "lojistik")),
):
    veri = RaporService.get_abc_analiz(db)
    RaporService.log_yaz(db, {"tur": "abc"}, "Basarili", current_user.id)
    return {"veri": veri}


@router.get("/doluluk")
def depo_doluluk_raporu(
    db: Session = Depends(get_db),
    current_user: Kullanici = Depends(require_role("admin", "lojistik", "depocu")),
):
    return {"veri": RaporService.get_doluluk(db)}


@router.post("/export")
def rapor_export(
    sablon_id: Optional[int] = None,
    tur: str = Query("stok", description="stok | siparis | hareket | kritik_stok | skt"),
    format: str = Query("csv", description="csv | excel"),
    baslang_tarihi: Optional[date] = Query(None),
    bitis_tarihi: Optional[date] = Query(None),
    db: Session = Depends(get_db),
    current_user: Kullanici = Depends(require_role("admin", "lojistik")),
):
    """Raporu CSV veya Excel olarak dışa aktar"""
    # raises BadRequestError for unknown tur
    rows = RaporService.get_export_data(db, tur, baslang_tarihi, bitis_tarihi)

    if not rows:
        from core.exceptions import NotFoundError
        raise NotFoundError("Rapor verisi", tur)

    RaporService.log_yaz(
        db, {"tur": tur, "format": format}, "Basarili", current_user.id, sablon_id=sablon_id
    )

    if format == "csv":
        output = io.StringIO()
        row_dicts = [dict(r._mapping) for r in rows] if hasattr(rows[0], "_mapping") else rows
        writer = csv.DictWriter(output, fieldnames=row_dicts[0].keys())
        writer.writeheader()
        writer.writerows(row_dicts)
        output.seek(0)
        return StreamingResponse(
            io.BytesIO(output.getvalue().encode("utf-8-sig")),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=rapor_{tur}.csv"},
        )

    if format == "excel":
        try:
            import openpyxl
            from openpyxl.styles import Font
        except ImportError:
            raise HTTPException(status_code=500, detail="Excel export için openpyxl paketi kurulu değil")

        row_dicts = [dict(r._mapping) for r in rows] if hasattr(rows[0], "_mapping") else rows
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = tur.capitalize()

        headers = list(row_dicts[0].keys())
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=str(header))
            cell.font = Font(bold=True)

        for row_idx, row_data in enumerate(row_dicts, 2):
            for col_idx, value in enumerate(row_data.values(), 1):
                ws.cell(row=row_idx, column=col_idx, value=str(value) if value is not None else "")

        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        return StreamingResponse(
            excel_buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=rapor_{tur}.xlsx"},
        )

    from core.exceptions import BadRequestError
    raise BadRequestError("Geçersiz format. 'csv' veya 'excel' kullanın.")


@router.post("/schedule/tetikle")
def rapor_schedule_tetikle(
    schedule_id: int,
    db: Session = Depends(get_db),
    current_user: Kullanici = Depends(require_role("admin")),
):
    return RaporService.tetikle_schedule(db, schedule_id, kullanici_id=current_user.id)


# ========================
# RAPOR LOGLARI
# ========================

@router.get("/log", response_model=list[RaporLoguResponse])
def rapor_loglari_listele(
    skip: int = 0,
    limit: int = 100,
    sablon_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    current_user: Kullanici = Depends(require_role("admin")),
):
    return RaporService.get_loglar(db, skip=skip, limit=limit, sablon_id=sablon_id)


# ========================
# RAPOR ZAMANLAMASI
# ========================

@router.get("/schedule", response_model=list[RaporScheduleResponse])
def rapor_schedules_listele(
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db),
    current_user: Kullanici = Depends(require_role("admin")),
):
    return RaporService.get_schedules(db, skip=skip, limit=limit)


@router.get("/schedule/{schedule_id}", response_model=RaporScheduleResponse)
def rapor_schedule_detay(
    schedule_id: int,
    db: Session = Depends(get_db),
    current_user: Kullanici = Depends(require_role("admin")),
):
    return RaporService.get_schedule(db, schedule_id)


@router.post("/schedule", response_model=RaporScheduleResponse, status_code=201)
def rapor_schedule_ekle(
    schedule: RaporScheduleCreate,
    db: Session = Depends(get_db),
    current_user: Kullanici = Depends(require_role("admin")),
):
    return RaporService.create_schedule(db, schedule, kullanici_id=current_user.id)


@router.put("/schedule/{schedule_id}", response_model=RaporScheduleResponse)
def rapor_schedule_guncelle(
    schedule_id: int,
    schedule_update: RaporScheduleUpdate,
    db: Session = Depends(get_db),
    current_user: Kullanici = Depends(require_role("admin")),
):
    return RaporService.update_schedule(db, schedule_id, schedule_update, kullanici_id=current_user.id)


@router.delete("/schedule/{schedule_id}")
def rapor_schedule_sil(
    schedule_id: int,
    db: Session = Depends(get_db),
    current_user: Kullanici = Depends(require_role("admin")),
):
    RaporService.delete_schedule(db, schedule_id, kullanici_id=current_user.id)
    return {"message": "Zamanlı rapor başarıyla silindi", "id": schedule_id}
